import openpyxl as xl

class Excel_Handler:

    def __init__(self, file_path):
        self.file_path = file_path

    def write_excel_dict(self, data):
        if not data:
            print("No data to write!")
            return  # Avoid errors when data is empty

        all_columns = list(data[0].keys())  # Extract column names

        # Create a new workbook
        workbook = xl.Workbook()
        sheet = workbook.active

        # Write headers
        sheet.append(all_columns)

        # Write data rows
        for row_dict in data:
            row_values = []
            for key in all_columns:  # Maintain column order
                row_values.append(row_dict.get(key, ""))  # Get value or empty string if missing
            sheet.append(row_values)

        # Save the workbook
        workbook.save(self.file_path)
        print(f"Excel file saved successfully at: {self.file_path}")

    def read_excel_dict(self):
        workbook = xl.load_workbook(self.file_path)
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:  # Assuming first row contains headers
            headers.append(cell.value)

        data_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            row_dict = {}
            for i in range(len(headers)):
                if row[i] is not None:
                    row_dict[headers[i]] = str(row[i])
                else:
                    row_dict[headers[i]] = ""
            data_list.append(row_dict)

        return data_list

if __name__ == '__main__':
    file_path = 'transactions.xlsx'
    excel_handler = Excel_Handler("Sample_Files\\DummyData.xlsx")
    data = excel_handler.read_excel_dict()
    excel_handler.write_excel_dict(data)